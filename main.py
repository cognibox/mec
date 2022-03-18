import argparse
import pytz
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import yaml

CBX_HEADER_LENGTH = 5
# noinspection SpellCheckingInspection
CBX_ID, CBX_COMPANY, CBX_CREATED_AT, CBX_USERNAME, CBX_OBJECT_CHANGES = range(CBX_HEADER_LENGTH)

# noinspection SpellCheckingInspection
cbx_headers = ['id', 'name', 'created_at', 'username', 'object_changes']

RESULT_HEADER_LENGTHS = 7
RES_ID, RES_CONTRACTOR, RES_USERNAME, RES_EXPIRATION_CHANGE_DATE, RES_EXPIRATION_PRIOR, RES_EXPIRATION_NEW,\
    RES_MEMBERSHIP_TYPE = range(RESULT_HEADER_LENGTHS)
res_headers = ['id', 'contractor_name', 'username', 'expiration_change_date', 'prior_expiration_date',
               'new_expiration_date', 'membership type']


# noinspection PyShadowingNames
def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


cbx_headers_text = '\n'.join([', '.join(x) for x in list(chunks(cbx_headers, 5))])
if len(cbx_headers) != CBX_HEADER_LENGTH:
    raise AssertionError('cbx header inconsistencies')

result_headers_text = '\n'.join([', '.join(x) for x in list(chunks(res_headers, 5))])
if len(res_headers) != RESULT_HEADER_LENGTHS:
    raise AssertionError('result header inconsistencies')


# define commandline parser
parser = argparse.ArgumentParser(
    description='Tool to analyse changes to expiration dates of memberships',
    formatter_class=argparse.RawTextHelpFormatter)
parser.add_argument('cbx_data',
                    help=f'xlsx DB export file of business units changes:\n{cbx_headers_text}\n\n')

parser.add_argument('output',
                    help=f'the xlsx file to be created with the analysis results:'
                         f'\n{result_headers_text}\n\n**Please note that metadata columns from the'
                         f' hc file are moved after the analysis data')

parser.add_argument('--for_month', dest='for_month', action='store',
                    help='to indicate what month the analysis is for in the following format: yyyy-mm')


parser.add_argument('--no_headers', dest='no_headers', action='store_true',
                    help='to indicate that input files have no headers')

parser.add_argument('--ignore_warnings', dest='ignore_warnings', action='store_true',
                    help='to ignore data consistency checks and run anyway...')

args = parser.parse_args()


# noinspection PyShadowingNames
def check_headers(headers, standards, ignore):
    headers = [x.lower().strip() for x in headers]
    for idx, val in enumerate(standards):
        if val != headers[idx]:
            print(f'WARNING: got "{headers[idx]}" while expecting "{val}" in column {idx + 1}')
            if not ignore:
                exit(-1)


if __name__ == '__main__':
    data_path = './data/'
    cbx_file = data_path + args.cbx_data
    output_file = data_path + args.output
    for_month = datetime.strptime(args.for_month, '%Y-%m') if args.for_month else None
    # output parameters used
    print(f'Starting at {datetime.now()}')
    print(f'Reading CBX data: {args.cbx_data}')
    print(f'analysing for {for_month.strftime("%b %Y") if for_month else "--"}')
    print(f'Outputting results in: {args.output}')
    # read data
    cbx_data = []
    print('Reading Cognibox data file...')
    cbx_wb = openpyxl.load_workbook(cbx_file, read_only=True)
    cbx_sheet = cbx_wb.active
    for row in cbx_sheet.rows:
        if not row[0].value:
            continue
        cbx_data.append([cell.value if cell.value else '' for cell in row])
    # check cbx db ata consistency
    if cbx_data and len(cbx_data[0]) != len(cbx_headers):
        print(f'WARNING: got {len(cbx_data[0])} columns when expecting {len(cbx_headers)}')
        if not args.ignore_warnings:
            exit(-1)
    if not args.no_headers:
        headers = cbx_data.pop(0)
        headers = [x.lower().strip() for x in headers]
        check_headers(headers, cbx_headers, args.ignore_warnings)

    # parse
    results = []
    prior_id = None
    row_item = []
    for row in cbx_data:
        if prior_id == row[CBX_ID]:
            append = True
        else:
            append = False
            if row_item and row_item[RES_EXPIRATION_NEW]:
                results.append(row_item)
            row_item = [None] * RESULT_HEADER_LENGTHS
            prior_id = row[CBX_ID]
        data = row[CBX_OBJECT_CHANGES].split('\n', 1)[1:][0]
        try:
            parsed_data = yaml.safe_load(data)
        except yaml.YAMLError as exc:
            print(exc)
            parsed_data = None

        for key, value in parsed_data.items():
            if key == 'cbx_expiration_date':
                prior = value[0] if value[0] is not None else ''
                new = value[1]
                row_item[RES_EXPIRATION_NEW] = new
                # correct the timestamp that is in UTC
                created = datetime.strptime(row[CBX_CREATED_AT], '%Y-%m-%dT%H:%M:%S.%f').replace(tzinfo=pytz.utc)
                created = created.astimezone(pytz.timezone('America/Montreal')).replace(tzinfo=None)
                row_item[RES_EXPIRATION_CHANGE_DATE] = created
                if row_item[RES_EXPIRATION_PRIOR] is None:
                    row_item[RES_EXPIRATION_PRIOR] = prior
        if not append:
            row_item[RES_USERNAME] = row[CBX_USERNAME]
            row_item[RES_ID] = row[CBX_ID]
            row_item[RES_CONTRACTOR] = row[CBX_COMPANY]

        start = for_month
    results = list(filter(lambda x: x[RES_EXPIRATION_CHANGE_DATE].month == for_month.month
                   and x[RES_EXPIRATION_CHANGE_DATE].year == for_month.year,
                          results))
    for result in results:
        if not result[RES_EXPIRATION_NEW] or type(result[RES_EXPIRATION_NEW]) is datetime:
            exp_new = result[RES_EXPIRATION_NEW]
        else:
            # noinspection PyTypeChecker
            exp_new = datetime.combine(result[RES_EXPIRATION_NEW], datetime.min.time())

        if not result[RES_EXPIRATION_PRIOR] or type(result[RES_EXPIRATION_PRIOR]) is datetime:
            exp_prior = result[RES_EXPIRATION_PRIOR]
        else:
            # noinspection PyTypeChecker
            exp_prior = datetime.combine(result[RES_EXPIRATION_PRIOR], datetime.min.time())
        if exp_new >= for_month + timedelta(days=365 * 21/12):
            result[RES_MEMBERSHIP_TYPE] = 'free'
        elif exp_new < for_month + timedelta(days=365 * 3/12):
            result[RES_MEMBERSHIP_TYPE] = 'expiration_change'
        elif not exp_prior or exp_prior + timedelta(days=365 * 6/12) < for_month:
            result[RES_MEMBERSHIP_TYPE] = 'new'
        else:
            result[RES_MEMBERSHIP_TYPE] = 'renewal'
    # write to excel
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = 'all'

    for index, value in enumerate(res_headers):
        out_ws.cell(1, index + 1, value)
    for index, result in enumerate(results):
        for i, value in enumerate(result):
            out_ws.cell(index + 2, i + 1, value)
    # format excel
    # formatting the excel...
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    dims = {}
    tab = Table(displayName=out_ws.title.replace(" ", "_"),
                ref=f'A1:{get_column_letter(out_ws.max_column)}{out_ws.max_row + 1}')
    tab.tableStyleInfo = style
    for row in out_ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        out_ws.column_dimensions[col].width = value
    out_ws.add_table(tab)
    out_wb.save(filename=output_file)
